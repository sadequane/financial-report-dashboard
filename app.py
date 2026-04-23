"""
Financial Report Dashboard — Streamlit App
-------------------------------------------
Author      : Portfolio Project — Sadequane
Background  : Culmination of "Python for Accountants" (Parts 1, 2 & 3)
Description : Applies concepts of DataFrames (Pandas), conditional logic, 
              loops, data structuring, and Excel automation to streamline 
              real-world accounting workflows.
Run         : streamlit run app.py
"""

import io
import warnings
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

warnings.filterwarnings("ignore")

# ── Page config ────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Financial Report Generator",
    page_icon="📊",
    layout="wide",
)

# ── Custom CSS ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main .block-container { padding-top: 2rem; }
    .metric-card {
        background: #f8f9fa;
        border-radius: 10px;
        padding: 1rem 1.2rem;
        border-left: 4px solid #2E75B6;
    }
    .stDownloadButton > button {
        background-color: #1E8449;
        color: white;
        border: none;
        border-radius: 8px;
        padding: 0.6rem 1.5rem;
        font-weight: 600;
    }
    .stDownloadButton > button:hover { background-color: #196F3D; }
    div[data-testid="metric-container"] {
        background: #f8f9fa;
        border-radius: 10px;
        padding: 1rem;
        border: 1px solid #e9ecef;
    }
</style>
""", unsafe_allow_html=True)


# ── Colour constants (Excel) ───────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
LIGHT_GREY = "F2F2F2"
GREEN      = "1E8449"


# ══════════════════════════════════════════════════════════════════════════
#  DATA PROCESSING (Applying concepts from Python for Accountants Part 2 & 3)
# ══════════════════════════════════════════════════════════════════════════
def load_and_clean(uploaded_file):
    """Read CSV, fix common messiness, return clean DataFrame."""
    df = pd.read_csv(uploaded_file)
    df.columns = [c.strip().title() for c in df.columns]

    required = {"Date", "Description", "Category", "Amount", "Type"}
    missing  = required - set(df.columns)
    if missing:
        st.error(f"CSV is missing columns: {missing}")
        st.stop()

    df["Date"]        = pd.to_datetime(df["Date"], format="mixed", dayfirst=False)
    df["Month"]       = df["Date"].dt.strftime("%b %Y")
    df["Month_Num"]   = df["Date"].dt.to_period("M")
    df["Description"] = df["Description"].str.strip().str.title()
    df["Category"]    = df["Category"].str.strip().str.title()
    df["Type"]        = df["Type"].str.strip().str.title()

    mask = df["Category"].isna() | (df["Category"] == "")
    df.loc[mask & (df["Type"] == "Income"), "Category"] = "Income"

    df["Amount"]      = pd.to_numeric(df["Amount"], errors="coerce").fillna(0)
    df["Income_Amt"]  = df["Amount"].where(df["Type"] == "Income",  0)
    df["Expense_Amt"] = df["Amount"].where(df["Type"] == "Expense", 0)

    return df.sort_values("Date").reset_index(drop=True)


def monthly_summary(df):
    grp = df.groupby("Month_Num").agg(
        Month        =("Month",       "first"),
        Total_Income =("Income_Amt",  "sum"),
        Total_Expenses=("Expense_Amt","sum"),
    ).reset_index(drop=True)
    grp["Net_Profit"]      = grp["Total_Income"] - grp["Total_Expenses"]
    grp["Profit_Margin_%"] = (
        grp["Net_Profit"] / grp["Total_Income"].replace(0, pd.NA) * 100
    ).round(1)
    return grp


def category_summary(df):
    exp = df[df["Type"] == "Expense"]
    grp = exp.groupby("Category")["Amount"].sum().reset_index()
    grp.columns = ["Category", "Total_Spent"]
    grp = grp.sort_values("Total_Spent", ascending=False).reset_index(drop=True)
    total = grp["Total_Spent"].sum()
    grp["% of Total"] = ((grp["Total_Spent"] / total) * 100).round(1)
    return grp


# ══════════════════════════════════════════════════════════════════════════
#  EXCEL REPORT BUILDER (Automating Excel Workpapers)
# ══════════════════════════════════════════════════════════════════════════
def _thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(ws, row, cols, col_start=1, bg=DARK_BLUE):
    for i, col in enumerate(cols):
        c = ws.cell(row=row, column=col_start + i, value=col)
        c.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _thin()


def build_excel(df, monthly, category):
    wb = Workbook()

    # ── Sheet 1: Dashboard ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"] = "Financial Report — Automated Excel Generator"
    ws["A1"].font      = Font(name="Arial", bold=True, size=16, color=WHITE)
    ws["A1"].fill      = PatternFill("solid", fgColor=DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    period = f"{df['Date'].min().strftime('%d %b %Y')}  –  {df['Date'].max().strftime('%d %b %Y')}"
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Period: {period}   |   Generated automatically"
    ws["A2"].font      = Font(name="Arial", italic=True, size=10, color="666666")
    ws["A2"].fill      = PatternFill("solid", fgColor="D6E4F0")
    ws["A2"].alignment = Alignment(horizontal="center")

    # KPIs
    total_inc = df["Income_Amt"].sum()
    total_exp = df["Expense_Amt"].sum()
    net       = total_inc - total_exp
    margin    = (net / total_inc * 100) if total_inc else 0

    kpis = [
        ("Total Income",   f"£{total_inc:,.2f}", "D5F5E3", "1E8449"),
        ("Total Expenses", f"£{total_exp:,.2f}", "FADBD8", "C0392B"),
        ("Net Profit",     f"£{net:,.2f}",       "D6E4F0", MID_BLUE),
        ("Profit Margin",  f"{margin:.1f}%",      "F2F2F2", DARK_BLUE),
    ]
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 32

    for idx, (label, value, bg, fg) in enumerate(kpis):
        col = 1 + idx * 2
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
        lc = ws.cell(row=4, column=col, value=label)
        lc.font      = Font(name="Arial", bold=True, size=9, color="666666")
        lc.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
        vc = ws.cell(row=5, column=col, value=value)
        vc.font      = Font(name="Arial", bold=True, size=14, color=fg)
        vc.fill      = PatternFill("solid", fgColor=bg)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border    = _thin()

    # Monthly table
    ws["A8"] = "Monthly Summary"
    ws["A8"].font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)
    _hdr(ws, 9, ["Month", "Income (£)", "Expenses (£)", "Net Profit (£)", "Margin (%)"])

    for r, row in monthly.iterrows():
        er    = 10 + r
        shade = LIGHT_GREY if r % 2 == 0 else WHITE
        vals  = [row["Month"], row["Total_Income"], row["Total_Expenses"],
                 row["Net_Profit"], row["Profit_Margin_%"]]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=er, column=c, value=val)
            cell.fill      = PatternFill("solid", fgColor=shade)
            cell.border    = _thin()
            cell.alignment = Alignment(horizontal="right" if c > 1 else "left")
            cell.font      = Font(name="Arial", size=10)
            if c in (2, 3, 4): cell.number_format = "#,##0.00"
            if c == 5:          cell.number_format = '0.0"%"'
            if c == 4:
                cell.font = Font(name="Arial", bold=True, size=10,
                                 color=GREEN if val >= 0 else "C0392B")

    tr = 10 + len(monthly)
    ws.cell(row=tr, column=1, value="TOTAL").fill = PatternFill("solid", fgColor="D6E4F0")
    ws.cell(row=tr, column=1).font  = Font(name="Arial", bold=True, size=10, color=DARK_BLUE)
    for c, col_l in [(2,"B"), (3,"C"), (4,"D")]:
        cell = ws.cell(row=tr, column=c, value=f"=SUM({col_l}10:{col_l}{tr-1})")
        cell.font          = Font(name="Arial", bold=True, size=10, color=DARK_BLUE)
        cell.fill          = PatternFill("solid", fgColor="D6E4F0")
        cell.border        = _thin()
        cell.number_format = "#,##0.00"
        cell.alignment     = Alignment(horizontal="right")

    for col, w in {"A":16,"B":15,"C":15,"D":16,"E":12,"F":15,"G":15}.items():
        ws.column_dimensions[col].width = w

    # ── Sheet 2: All Transactions ───────────────────────────────────────
    ws2 = wb.create_sheet("All Transactions")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:F1")
    ws2["A1"] = "All Transactions — Cleaned Data"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws2["A1"].fill      = PatternFill("solid", fgColor=MID_BLUE)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28
    _hdr(ws2, 2, ["Date","Description","Category","Type","Amount (£)","Month"], bg=MID_BLUE)

    for r, row in df.iterrows():
        er    = 3 + r
        shade = LIGHT_GREY if r % 2 == 0 else WHITE
        vals  = [row["Date"].strftime("%d/%m/%Y"), row["Description"],
                 row["Category"], row["Type"], row["Amount"], row["Month"]]
        for c, val in enumerate(vals, 1):
            cell = ws2.cell(row=er, column=c, value=val)
            cell.fill      = PatternFill("solid", fgColor=shade)
            cell.border    = _thin()
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="right" if c == 5 else "left")
            if c == 5: cell.number_format = "#,##0.00"
            if row["Type"] == "Income"  and c == 4:
                cell.font = Font(name="Arial", bold=True, size=10, color=GREEN)
            if row["Type"] == "Expense" and c == 4:
                cell.font = Font(name="Arial", bold=True, size=10, color="C0392B")

    for col, w in {"A":13,"B":38,"C":16,"D":10,"E":13,"F":12}.items():
        ws2.column_dimensions[col].width = w

    # ── Sheet 3: Expense Breakdown ──────────────────────────────────────
    ws3 = wb.create_sheet("Expense Breakdown")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:C1")
    ws3["A1"] = "Expense Breakdown by Category"
    ws3["A1"].font      = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws3["A1"].fill      = PatternFill("solid", fgColor=DARK_BLUE)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28
    _hdr(ws3, 2, ["Category","Total Spent (£)","% of Total"])

    for r, row in category.iterrows():
        er    = 3 + r
        shade = LIGHT_GREY if r % 2 == 0 else WHITE
        for c, val in enumerate([row["Category"], row["Total_Spent"], row["% of Total"]], 1):
            cell = ws3.cell(row=er, column=c, value=val)
            cell.fill      = PatternFill("solid", fgColor=shade)
            cell.border    = _thin()
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="right" if c > 1 else "left")
            if c == 2: cell.number_format = "#,##0.00"
            if c == 3: cell.number_format = '0.0"%"'

    tr3 = 3 + len(category)
    ws3.cell(row=tr3, column=1, value="TOTAL").fill = PatternFill("solid", fgColor="D6E4F0")
    ws3.cell(row=tr3, column=1).font = Font(name="Arial", bold=True, size=10, color=DARK_BLUE)
    t = ws3.cell(row=tr3, column=2, value=f"=SUM(B3:B{tr3-1})")
    t.font = Font(name="Arial", bold=True, size=10, color=DARK_BLUE)
    t.fill = PatternFill("solid", fgColor="D6E4F0"); t.border = _thin()
    t.number_format = "#,##0.00"; t.alignment = Alignment(horizontal="right")

    chart = BarChart()
    chart.type  = "col"
    chart.title = "Expenses by Category"
    chart.y_axis.title = "Amount (£)"
    chart.style = 10; chart.width = 18; chart.height = 12
    data = Reference(ws3, min_col=2, min_row=2, max_row=2+len(category))
    cats = Reference(ws3, min_col=1, min_row=3, max_row=2+len(category))
    chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
    ws3.add_chart(chart, "E3")
    for col, w in {"A":20,"B":16,"C":14}.items():
        ws3.column_dimensions[col].width = w

    # Save to bytes buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════

# ── Sidebar Portfolio Context ──────────────────────────────────────────────
with st.sidebar:
    st.header("About This Project")
    st.markdown("""
    **Portfolio Project by Sadequane**
    
    This app was built as a capstone project after completing the **Python for Accountants** series (Parts 1, 2 & 3). 
    
    It demonstrates the practical application of programming in real-life accounting workflows by:
    * Reading and processing CSV files.
    * Utilizing **Pandas DataFrames** to calculate Net Profit and profit margins.
    * Automating repetitive tasks with loops and conditional statements.
    * Replacing manual Excel work by automatically generating formatted `.xlsx` workpapers.
    """)

st.title("📊 Financial Report Generator")
st.markdown("Upload your `transactions.csv` and get a live dashboard plus a formatted Excel report instantly.")

uploaded = st.file_uploader(
    "Drop your CSV file here",
    type=["csv"],
    help="CSV must have columns: Date, Description, Category, Amount, Type"
)

# Show sample format hint
with st.expander("📋 What format should my CSV be in?"):
    st.markdown("Your CSV needs these 5 columns:")
    st.code("Date, Description, Category, Amount, Type", language="text")
    sample = pd.DataFrame({
        "Date":        ["2024-01-03", "2024-01-05"],
        "Description": ["Office Supplies", "Client Payment - Acme Ltd"],
        "Category":    ["Office", "Income"],
        "Amount":      [120.50, 5000.00],
        "Type":        ["Expense", "Income"],
    })
    st.dataframe(sample, hide_index=True, use_container_width=True)

# ── Main content (only shows after upload) ─────────────────────────────────
if uploaded:
    df       = load_and_clean(uploaded)
    monthly  = monthly_summary(df)
    category = category_summary(df)

    total_inc = df["Income_Amt"].sum()
    total_exp = df["Expense_Amt"].sum()
    net       = total_inc - total_exp
    margin    = (net / total_inc * 100) if total_inc else 0

    st.success(f"✅ {len(df)} transactions loaded and cleaned successfully.")

    # ── KPI metrics ──────────────────────────────────────────────────────
    st.markdown("### Summary")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Income",   f"£{total_inc:,.2f}")
    c2.metric("Total Expenses", f"£{total_exp:,.2f}")
    c3.metric("Net Profit",     f"£{net:,.2f}")
    c4.metric("Profit Margin",  f"{margin:.1f}%")

    st.markdown("---")

    # ── Monthly summary table ────────────────────────────────────────────
    col_left, col_right = st.columns(2)

    with col_left:
        st.markdown("### Monthly Summary")
        display_monthly = monthly[["Month","Total_Income","Total_Expenses","Net_Profit","Profit_Margin_%"]].copy()
        display_monthly.columns = ["Month","Income (£)","Expenses (£)","Net Profit (£)","Margin (%)"]
        st.dataframe(
            display_monthly.style.format({
                "Income (£)"    : "£{:,.2f}",
                "Expenses (£)"  : "£{:,.2f}",
                "Net Profit (£)": "£{:,.2f}",
                "Margin (%)"    : "{:.1f}%",
            }).map(
                lambda v: "color: green; font-weight: bold" if isinstance(v, float) and v > 0 else
                          "color: red;   font-weight: bold" if isinstance(v, float) and v < 0 else "",
            ),
            use_container_width=True,
            hide_index=True,
        )

    with col_right:
        st.markdown("### Expenses by Category")
        display_cat = category.copy()
        display_cat.columns = ["Category","Total Spent (£)","% of Total"]
        st.dataframe(
            display_cat.style.format({
                "Total Spent (£)": "£{:,.2f}",
                "% of Total"     : "{:.1f}%",
            }),
            use_container_width=True,
            hide_index=True,
        )

    st.markdown("---")

    # ── Recent transactions ──────────────────────────────────────────────
    st.markdown("### All Transactions")
    display_tx = df[["Date","Description","Category","Type","Amount"]].copy()
    display_tx["Date"]   = display_tx["Date"].dt.strftime("%d/%m/%Y")
    display_tx["Amount"] = display_tx["Amount"].map("£{:,.2f}".format)
    st.dataframe(display_tx, use_container_width=True, hide_index=True)

    st.markdown("---")

    # ── Download button ──────────────────────────────────────────────────
    st.markdown("### Download Your Report")
    excel_buf = build_excel(df, monthly, category)

    st.download_button(
        label="⬇️ Download Excel Report (.xlsx)",
        data=excel_buf,
        file_name="Financial_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.caption("The report includes 3 sheets: Dashboard · All Transactions · Expense Breakdown")

else:
    st.info("👆 Upload a CSV file above to get started.")
